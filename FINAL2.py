import sys
import os
from PIL import Image, ImageDraw, ImageFont
import PIL
from glob import glob
import openpyxl
from PIL import Image, ImageDraw, ImageFont
abspath = os.path.dirname(sys.argv[0])
dname = os.path.dirname(abspath)


	
def CropImage(str):
	os.chdir(image_directory)
	temp = Image.open(str)
	width,height = temp.size
	temp1 = temp.resize((int(width/4),int(height/4)))
	temp2 = temp.resize((int(width/20),int(height/20)))
	os.chdir(pic_directory)
	shit = "z" + str
	temp1.save(shit)
	os.chdir(thumb_directory)
	temp2.save(shit)
	
def PopulateXcel():
	os.chdir(abspath)
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = 'Data for HTML'
	files_list = []
	files_list = glob(os.path.join(pic_directory, '*.jpg'))
	tot = len(files_list)
	num = 1
	ws.cell(row = num, column = 1).value = "alt"
	ws.cell(row = num, column = 2).value = "src"
	ws.cell(row = num, column = 3).value = "data-type"
	ws.cell(row = num, column = 4).value = "data-image"
	ws.cell(row = num, column = 5).value = "data-description"
	ws.cell(row = num, column = 6).value = "data-videoid"
	for a_file in sorted(files_list):
		alt = None
		src = None
		dtype = None
		dimage = None
		ddescription = None
		dvideoid = None
		b = a_file
		my = os.path.split(b)
		last = my[1]
		AddLogo(last)
		name = last
		my = os.path.split(my[0])
		first = my[1]
		pth = first+"/"+last
		dimage = pth
		src = "Thumbs"+"/"+last
		alt = last.rstrip('.jpg')
		ddescription = "This is an image of a --placeholder--"
		tot = tot + 2
		num = num + 1
		ws.cell(row = num, column = 1).value = alt
		ws.cell(row = num, column = 2).value = src
		ws.cell(row = num, column = 3).value = dtype
		ws.cell(row = num, column = 4).value = dimage
		ws.cell(row = num, column = 5).value = ddescription
		ws.cell(row = num, column = 6).value = dvideoid

	wb.save('PictureInfoForWeb.xlsx')
	
	
def AddLogo(curimage):
	im = Image.new('RGBA', (200, 200),'white')
	draw = ImageDraw.Draw(im)
	fontsFolder = 'FONT_FOLDER'
	Font = ImageFont.truetype(os.path.join("C:\\Windows\\Fonts\\", 'MATURASC.TTF'), 32) 
	str = "DC"
	draw.text((60, 80), str, fill='grey', font=Font)
	imd = im.crop((50,50,150,150))
	os.chdir(pic_directory)
	main = Image.open(curimage)
	width,height = main.size
	width = width - 150
	height = height - 150
	main.paste(imd,(width,height))
	os.chdir("D:\\WebTrial\\Pics")
	main.save(curimage)
	os.chdir(abspath)
	
def CodeTextMaker():
	wb = openpyxl.load_workbook(filename = 'PictureInfoForWeb.xlsx', read_only = True)
	ws = wb.get_active_sheet()
	titles = []
	string = "\t" + "<img"
	max = 0
	rcnt = 1
	ptr = 0
	txtfile = open("code.txt","w+")
	ctr = 0
	spacefive = (" ")*5
	spaceone = " "
	for row in ws.rows:
		for cell in row:
			if(rcnt == 1):
				titles.append(cell.value)
				max = len(titles)
				ctr = 1
			else:
				temp = cell.value
				if(ptr < max):
					if(temp == None):
						ptr = ptr + 1
						continue
					else:
						if(ptr == 0):
							string = string + spaceone + str(titles[ptr]) + "=" + "\"" + str(temp) + "\""
							ptr = ptr + 1
						else:
							string = string + "\n" + "\t" + spacefive + str(titles[ptr]) + "=" + "\"" + str(temp) + "\""
							ptr = ptr + 1
		if(ctr == 0):
			string = string + ">" + "\n" + "\n"
			txtfile.write(string)
		rcnt = rcnt + 1
		ptr = 0
		ctr = 0
		string = "\t" + "<img"
		

	txtfile.close()
	
def FinalHTML():
	main_file = open('combo.txt','w+')
	file1 = open('codes.txt','r+')
	file2 = open('code.txt','r+')
	file3= open('codesl.txt','r+')
	frst= file1.read()
	betw = file2.read()
	lst = file3.read()
	total = frst + "\n" + betw + "\n" + lst
	main_file.write(total)
	main_file.close()
	file1.close()
	file2.close()
	file3.close()
	f = open('combo.txt','r')
	filedata = f.read()
	f.close()
	newdata = filedata.replace("WebPageTitle",WebPageTitle)
	filedata = newdata
	newdata = filedata.replace("PageTitle",PageTitle)
	filedata = newdata
	newdata = filedata.replace("PageDescription",PageDescription)
	filedata = newdata
	f = open('combo.txt','w')
	f.write(newdata)
	f.close()
	os.chdir(web_directory)
	html_file = open('webpagenew.html','w+')
	html_file.write(newdata)
	html_file.close()

for i in range(10):
	image_directory = raw_input("Please enter the path of the image. Note: Take care NOT to use \\ or \"\" in the path. \n")
	print "You Entered: " + image_directory
	Dec = raw_input('\n Do you want to change Y/N? \n')
	if(Dec == 'Y' or Dec == 'y'):
		image_directory = raw_input("Please enter the path of the image. Note: Take care NOT to use \\ or \"\" in the path.\n")
	else:
		break
	dec = 'N'
for i in range(10):
	pic_directory = raw_input("Please enter the path where the cropped images must be stored. Note: Take care to use \\ in the path.\n")
	print "You Entered: " + pic_directory
	Dec = raw_input('\n Do you want to change Y/N? \n')
	if(Dec == 'Y' or Dec == 'y'):
		pic_directory = raw_input("Please enter the path where the cropped images must be stored. Note: Take care to use \\ in the path.")
	else:
		break
	dec = 'N'
for i in range(10):
	thumb_directory = raw_input("Please enter the path where the Thumbnails need to be stored. Note: Take care to use \\ in the path.\n")
	print "You Entered: " + thumb_directory
	Dec = raw_input('\n Do you want to change Y/N? \n')
	if(Dec == 'Y' or Dec == 'y'):
		thumb_directory = raw_input("\nPlease enter the path where the Thumbnails need to be stored. Note: Take care to use \\ in the path.\n")
	else:
		break
	dec = 'N'
for i in range(10):
	web_directory = raw_input("\nPlease enter where you want to store the webpage?\n")
	print "You Entered: " + web_directory
	Dec = raw_input('\nDo you want to change Y/N?\n')
	if(Dec == 'Y' or Dec == 'y'):
		web_directory = raw_input("Please enter where you want to store the webpage?\n")
	else:
		break
	dec = 'N'
	
WebPageTitle = raw_input('Give this webpage a name?')
PageTitle = raw_input('Give this webpage a title or What is this webpage about?')
PageDescription = raw_input('Write a few words describing this page')



os.chdir(image_directory)
files_list = glob(os.path.join(image_directory, '*.jpg'))
tot = len(files_list)
for b in files_list:
	my = os.path.split(b)
	last = my[1]
	CropImage(last)
	

PopulateXcel()
os.chdir(abspath)
CodeTextMaker()
FinalHTML()
os.chdir(abspath)
raw_input('Successful? Rate me out of 5')

